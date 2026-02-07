# 1️⃣ FastAPI y utilidades
from fastapi import APIRouter, Request, Depends, Form
from fastapi.responses import RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from typing import Optional

# 2️⃣ SQLAlchemy
from sqlalchemy.orm import Session
from sqlalchemy import func

# 3️⃣ Utilidades estándar
from datetime import datetime, timedelta
from io import BytesIO

# 4️⃣ App interna (DB primero)
from app.db import get_db

# 5️⃣ MODELOS (esto es CLAVE)
from app.models import Producto, Sucursal
from app.models.stock import Stock
from app.models.ventas import Venta, VentaItem

# 6️⃣ Extras (Excel)
from fastapi.responses import StreamingResponse
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import json

router = APIRouter(prefix="/admin", tags=["Admin"])
templates = Jinja2Templates(directory="app/templates")


@router.get("/")
def admin_home():
    return RedirectResponse(url="/admin/dashboard", status_code=302)


@router.get("/dashboard")
def admin_dashboard(
    request: Request,
    db: Session = Depends(get_db),
    sucursal_id: int | None = None,
    days: int = 7,
):
    # ---- Validaciones básicas ----
    if days not in (7, 14, 30, 60):
        days = 7

    # ---- Métricas generales ----
    total_productos = db.query(Producto).count()
    total_sucursales = db.query(Sucursal).count()
    total_registros_stock = db.query(Stock).count()

    LOW_STOCK = 5

    sin_stock = db.query(Stock).filter(Stock.cantidad == 0).count()

    bajo_stock = (
        db.query(Stock).filter(Stock.cantidad > 0, Stock.cantidad <= LOW_STOCK).count()
    )

    ok_stock = db.query(Stock).filter(Stock.cantidad > LOW_STOCK).count()

    # ---- Filtro por sucursal (opcional) ----
    ventas_query = db.query(Venta)

    if sucursal_id:
        ventas_query = ventas_query.filter(Venta.sucursal_id == sucursal_id)

    # ---- Gráfico: ventas por día ----
    start_date = (datetime.utcnow() - timedelta(days=days - 1)).date()

    rows = (
        ventas_query.filter(
            Venta.created_at >= datetime.utcnow() - timedelta(days=days)
        )
        .with_entities(
            func.date(Venta.created_at).label("d"), func.count(Venta.id).label("c")
        )
        .group_by(func.date(Venta.created_at))
        .order_by(func.date(Venta.created_at))
        .all()
    )

    mapa = {str(r.d): int(r.c) for r in rows}

    chart_labels: list[str] = []
    chart_values: list[int] = []

    for i in range(days):
        d = start_date + timedelta(days=i)
        key = d.isoformat()
        chart_labels.append(d.strftime("%a"))  # Lun, Mar, Mié...
        chart_values.append(mapa.get(key, 0))

    # ---- Sucursales para selector ----
    sucursales = db.query(Sucursal).order_by(Sucursal.nombre).all()

    # ---- Render ----
    return templates.TemplateResponse(
        "admin/dashboard.html",
        {
            "request": request,
            "total_productos": total_productos,
            "total_sucursales": total_sucursales,
            "total_registros_stock": total_registros_stock,
            "sin_stock": sin_stock,
            "bajo_stock": bajo_stock,
            "ok_stock": ok_stock,
            "chart_labels": chart_labels,
            "chart_values": chart_values,
            "days": days,
            "sucursales": sucursales,
            "sucursal_id": sucursal_id,
        },
    )


@router.get("/productos")
def admin_productos(request: Request, db: Session = Depends(get_db)):
    productos = db.query(Producto).order_by(Producto.id.desc()).all()

    return templates.TemplateResponse(
        "admin/productos.html",
        {
            "request": request,
            "title": "Productos",
            "subtitle": "Catálogo y precios",
            "productos": productos,
            "active": "productos",
        },
    )


@router.post("/productos/crear")
def admin_crear_producto(
    nombre: str = Form(...),
    sku: str = Form(""),
    precio: float = Form(0),
    descripcion: str = Form(""),
    db: Session = Depends(get_db),
):
    nombre = nombre.strip()
    sku = sku.strip()

    p = Producto(
        nombre=nombre,
        sku=sku if sku else f"SKU-{nombre[:10]}",
        precio=precio,
        descripcion=descripcion if descripcion else None,
    )
    db.add(p)
    db.commit()
    return RedirectResponse(url="/admin/productos", status_code=303)


@router.get("/stock")
def admin_stock(
    request: Request, sucursal_id: Optional[int] = None, db: Session = Depends(get_db)
):
    # Umbrales de stock
    STOCK_BAJO = 10
    STOCK_CRITICO = 5

    sucursales = db.query(Sucursal).order_by(Sucursal.id.asc()).all()
    if sucursal_id is None and sucursales:
        sucursal_id = sucursales[0].id

    productos = db.query(Producto).order_by(Producto.nombre.asc()).all()

    stock_map = {}
    if sucursal_id is not None:
        stocks = db.query(Stock).filter(Stock.sucursal_id == sucursal_id).all()
        stock_map = {st.producto_id: st.cantidad for st in stocks}

    return templates.TemplateResponse(
        "admin/stock.html",
        {
            "request": request,
            "productos": productos,
            "stock_map": stock_map,
            "sucursales": sucursales,
            "sucursal_id": sucursal_id,
            "STOCK_BAJO": STOCK_BAJO,
            "STOCK_CRITICO": STOCK_CRITICO,
            "active": "stock",
        },
    )


@router.post("/stock/ajustar")
def admin_ajustar_stock(
    sucursal_id: int = Form(...),
    producto_id: int = Form(...),
    delta: int = Form(...),
    db: Session = Depends(get_db),
):
    st = (
        db.query(Stock)
        .filter(Stock.sucursal_id == sucursal_id, Stock.producto_id == producto_id)
        .first()
    )

    if not st:
        st = Stock(sucursal_id=sucursal_id, producto_id=producto_id, cantidad=0)
        db.add(st)

    nueva = (st.cantidad or 0) + int(delta)
    if nueva < 0:
        nueva = 0

    st.cantidad = nueva
    db.commit()

    return RedirectResponse(
        url=f"/admin/stock?sucursal_id={sucursal_id}", status_code=303
    )


@router.get("/ventas")
def admin_ventas(request: Request, db: Session = Depends(get_db)):
    sucursales = db.query(Sucursal).order_by(Sucursal.id.asc()).all()
    productos = db.query(Producto).order_by(Producto.id.asc()).all()

    # últimas 20 ventas
    ventas = db.query(Venta).order_by(Venta.id.desc()).limit(20).all()

    return templates.TemplateResponse(
        "admin/ventas.html",
        {
            "request": request,
            "title": "Ventas",
            "subtitle": "Registrar ventas y descontar stock",
            "active": "ventas",
            "sucursales": sucursales,
            "productos": productos,
            "ventas": ventas,
        },
    )


@router.post("/ventas")
def crear_venta(
    request: Request,
    db: Session = Depends(get_db),
    sucursal_id: int = Form(...),
    producto_id: int = Form(...),
    qty: int = Form(...),
):
    if qty <= 0:
        return RedirectResponse(url="/admin/ventas", status_code=303)

    # buscar stock del producto en la sucursal
    stock = (
        db.query(Stock)
        .filter(Stock.sucursal_id == sucursal_id, Stock.producto_id == producto_id)
        .first()
    )
    if not stock:
        # si no existe, lo creas en 0
        stock = Stock(sucursal_id=sucursal_id, producto_id=producto_id, cantidad=0)
        db.add(stock)
        db.flush()

    # no permitir negativo
    if stock.cantidad < qty:
        return RedirectResponse(url="/admin/ventas?err=nostock", status_code=303)

    # crear venta + item
    venta = Venta(sucursal_id=sucursal_id)
    db.add(venta)
    db.flush()  # para obtener venta.id

    item = VentaItem(venta_id=venta.id, producto_id=producto_id, qty=qty, precio=0)
    db.add(item)

    # descontar stock
    stock.cantidad -= qty

    db.commit()
    return RedirectResponse(url="/admin/ventas?ok=1", status_code=303)


@router.get("/analisis")
def admin_analisis(
    request: Request,
    db: Session = Depends(get_db),
    days: int = 7,  # ?days=7 / 30 / 90
):
    # --- parámetros ---
    days = int(days) if str(days).isdigit() else 7
    if days not in (7, 30, 90):
        days = 7

    now = datetime.utcnow()
    start_cur = now - timedelta(days=days)
    start_prev = now - timedelta(days=days * 2)
    end_prev = start_cur

    # --- KPIs base ---
    total_ventas_cur = db.query(Venta).filter(Venta.created_at >= start_cur).count()

    unidades_cur = (
        db.query(func.coalesce(func.sum(VentaItem.qty), 0))
        .join(Venta, Venta.id == VentaItem.venta_id)
        .filter(Venta.created_at >= start_cur)
        .scalar()
    ) or 0

    ticket_prom_unidades = (
        round((unidades_cur / total_ventas_cur), 2) if total_ventas_cur else 0
    )

    # --- Gráfico: ventas por día (actual vs anterior) ---
    # ventas por día (conteo de ventas)
    rows_cur = (
        db.query(
            func.date(Venta.created_at).label("d"), func.count(Venta.id).label("c")
        )
        .filter(Venta.created_at >= start_cur)
        .group_by(func.date(Venta.created_at))
        .order_by(func.date(Venta.created_at))
        .all()
    )
    rows_prev = (
        db.query(
            func.date(Venta.created_at).label("d"), func.count(Venta.id).label("c")
        )
        .filter(Venta.created_at >= start_prev, Venta.created_at < end_prev)
        .group_by(func.date(Venta.created_at))
        .order_by(func.date(Venta.created_at))
        .all()
    )

    map_cur = {str(r.d): int(r.c) for r in rows_cur}
    map_prev = {str(r.d): int(r.c) for r in rows_prev}

    # labels fijos (últimos N días)
    start_date = (now - timedelta(days=days - 1)).date()
    labels = []
    values_cur = []
    values_prev = []

    for i in range(days):
        d = start_date + timedelta(days=i)
        key = d.isoformat()
        labels.append(d.strftime("%a"))  # Lun, Mar... (depende locale del sistema)
        values_cur.append(map_cur.get(key, 0))

        # para el periodo anterior, alineamos por “día equivalente”
        d_prev = d - timedelta(days=days)
        key_prev = d_prev.isoformat()
        values_prev.append(map_prev.get(key_prev, 0))

    # --- Donut: ventas por sucursal (conteo) ---
    suc_rows = (
        db.query(Sucursal.nombre, func.count(Venta.id))
        .join(Venta, Venta.sucursal_id == Sucursal.id)
        .filter(Venta.created_at >= start_cur)
        .group_by(Sucursal.nombre)
        .order_by(func.count(Venta.id).desc())
        .all()
    )
    donut_labels = [r[0] for r in suc_rows]
    donut_values = [int(r[1]) for r in suc_rows]

    # --- Top productos (unidades) ---
    top_prod = (
        db.query(Producto.nombre, func.coalesce(func.sum(VentaItem.qty), 0).label("u"))
        .join(VentaItem, VentaItem.producto_id == Producto.id)
        .join(Venta, Venta.id == VentaItem.venta_id)
        .filter(Venta.created_at >= start_cur)
        .group_by(Producto.nombre)
        .order_by(func.coalesce(func.sum(VentaItem.qty), 0).desc())
        .limit(7)
        .all()
    )
    top_labels = [r[0] for r in top_prod]
    top_values = [int(r[1]) for r in top_prod]

    return templates.TemplateResponse(
        "admin/analisis.html",
        {
            "request": request,
            "title": "Análisis",
            "subtitle": f"Indicadores ({days} días)",
            "active": "analisis",
            "days": days,
            "total_ventas_cur": total_ventas_cur,
            "unidades_cur": int(unidades_cur),
            "ticket_prom_unidades": ticket_prom_unidades,
            # charts (mandamos listas, NO json.dumps)
            "labels": labels,
            "values_cur": values_cur,
            "values_prev": values_prev,
            "donut_labels": donut_labels,
            "donut_values": donut_values,
            "top_labels": top_labels,
            "top_values": top_values,
        },
    )


@router.get("/reportes")
def admin_reportes(request: Request):
    return templates.TemplateResponse(
        "admin/reportes.html",
        {
            "request": request,
            "title": "Reportes Excel",
            "subtitle": "Descarga reportes en formato .xlsx",
            "active": "reportes",
        },
    )


@router.get("/reportes/ventas.xlsx")
def reporte_ventas_excel(db: Session = Depends(get_db)):
    # Trae ventas (ajusta el límite si quieres)
    ventas = db.query(Venta).order_by(Venta.id.desc()).limit(1000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    # Encabezados
    ws.append(["Venta ID", "Fecha", "Sucursal ID", "Producto ID", "Cantidad", "Precio"])

    # Filas
    for v in ventas:
        for it in v.items:
            ws.append(
                [
                    v.id,
                    v.created_at.strftime("%Y-%m-%d %H:%M:%S") if v.created_at else "",
                    v.sucursal_id,
                    it.producto_id,
                    it.qty,
                    it.precio,
                ]
            )

    # Auto-ancho columnas (simple)
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="ventas.xlsx"'},
    )
